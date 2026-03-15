from datetime import date
from typing import Dict, Any, List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_

from app import models

import logging
logger = logging.getLogger(__name__)

class ReportingService:
    """
    Financial & Regulatory Reporting Engine (COBAC Compliant)
    Relies exclusively on the Double-Entry General Ledger (GLJournalEntry) for accuracy.
    """

    @staticmethod
    def generate_trial_balance(db: Session, as_of_date: date) -> List[Dict[str, Any]]:
        """
        Generates the Trial Balance (Balance Générale).
        Returns a list of all GL accounts with their total Debits, Credits, and Net Balance.
        Must balance to exactly 0 globally.
        """
        results = db.query(
            models.GLAccount.account_code,
            models.GLAccount.account_name,
            models.GLAccount.account_type,
            func.sum(
                case(
                    (func.date(models.GLJournalEntry.entry_date) < as_of_date, 
                     case((models.GLJournalEntry.entry_type == 'DEBIT', models.GLJournalEntry.amount), else_=-models.GLJournalEntry.amount)),
                    else_=0
                )
            ).label('opening_balance'),
            func.sum(
                case(
                    (and_(func.date(models.GLJournalEntry.entry_date) == as_of_date, models.GLJournalEntry.entry_type == 'DEBIT'), models.GLJournalEntry.amount),
                    else_=0
                )
            ).label('total_debit'),
            func.sum(
                case(
                    (and_(func.date(models.GLJournalEntry.entry_date) == as_of_date, models.GLJournalEntry.entry_type == 'CREDIT'), models.GLJournalEntry.amount),
                    else_=0
                )
            ).label('total_credit')
        ).outerjoin(
            models.GLJournalEntry, 
            models.GLAccount.id == models.GLJournalEntry.gl_account_id
        ).group_by(
            models.GLAccount.account_code,
            models.GLAccount.account_name,
            models.GLAccount.account_type
        ).order_by(models.GLAccount.account_code).all()

        trial_balance = []
        for row in results:
            opening = float(row.opening_balance or 0)
            debit = float(row.total_debit or 0)
            credit = float(row.total_credit or 0)
            
            # Closing Balance: Opening (signed) + today's Debit - today's Credit
            closing = opening + debit - credit

            # balance_type tells consumers whether this account has a net DEBIT or CREDIT balance
            balance_type = "DEBIT" if closing >= 0 else "CREDIT"
            
            # COBAC: Trial Balance should show the full structure (Audit trail)
            # Include if there is any movement or if it's a DETAIL account (leaf node)
            trial_balance.append({
                "account_code": row.account_code,
                "account_name": row.account_name,
                "account_type": row.account_type,
                "opening_balance": opening,
                "debit": debit,
                "credit": credit,
                "closing_balance": closing,
                # These fields are required by generate_balance_sheet & net income calc:
                "balance": abs(closing),
                "balance_type": balance_type
            })

        # Compute totals row (a valid trial balance has total_debit == total_credit)
        total_opening = sum(r["opening_balance"] for r in trial_balance)
        total_debit = sum(r["debit"] for r in trial_balance)
        total_credit = sum(r["credit"] for r in trial_balance)
        net_balance = total_opening + total_debit - total_credit  # Should be 0 if balanced

        return {
            "as_of_date": as_of_date.isoformat(),
            "rows": trial_balance,
            "totals": {
                "opening_balance": round(total_opening, 2),
                "debit": round(total_debit, 2),
                "credit": round(total_credit, 2),
                "net_balance": round(net_balance, 2),
                "is_balanced": abs(net_balance) < 0.01
            }
        }

    @staticmethod
    def generate_balance_sheet(db: Session, as_of_date: date) -> Dict[str, Any]:
        """
        Generates the Balance Sheet (Bilan).
        Assets = Liabilities + Equity
        """
        tb_result = ReportingService.generate_trial_balance(db, as_of_date)
        # generate_trial_balance now returns a dict with 'rows' and 'totals'
        tb = tb_result["rows"]
        
        assets = []
        liabilities = []
        equity = []
        
        total_assets = 0.0
        total_liabilities = 0.0
        total_equity = 0.0

        for account in tb:
            val = account['balance'] if account['balance_type'] == 'DEBIT' else -account['balance']
            if account['account_type'] == 'ASSET':
                assets.append(account)
                total_assets += account['balance'] if account['balance_type'] == 'DEBIT' else -account['balance']
            
            val = account['balance'] if account['balance_type'] == 'CREDIT' else -account['balance']
            if account['account_type'] == 'LIABILITY':
                liabilities.append(account)
                total_liabilities += account['balance'] if account['balance_type'] == 'CREDIT' else -account['balance']
            elif account['account_type'] == 'EQUITY':
                equity.append(account)
                total_equity += account['balance'] if account['balance_type'] == 'CREDIT' else -account['balance']

        # Note: Current Year Net Income should roll into Equity for the BS to balance perfectly.
        # We simulate this by calculating net income from the TB
        net_income = sum([a['balance'] if a['balance_type'] == 'CREDIT' else -a['balance'] for a in tb if a['account_type'] == 'INCOME']) - \
                     sum([a['balance'] if a['balance_type'] == 'DEBIT' else -a['balance'] for a in tb if a['account_type'] == 'EXPENSE'])
                     
        total_equity += net_income
        equity.append({
            "account_code": "NET_INC",
            "account_name": "Net Income (Current Period)",
            "account_type": "EQUITY",
            "balance": abs(net_income),
            "balance_type": "CREDIT" if net_income >=0 else "DEBIT"
        })

        return {
            "as_of_date": as_of_date,
            "assets": {"items": assets, "total": total_assets},
            "liabilities": {"items": liabilities, "total": total_liabilities},
            "equity": {"items": equity, "total": total_equity},
            "is_balanced": abs(total_assets - (total_liabilities + total_equity)) < 0.01 # Float precision allowance
        }

    @staticmethod
    def generate_income_statement(db: Session, start_date: date, end_date: date) -> Dict[str, Any]:
        """
        Generates the Income Statement (Compte de Résultat).
        Revenue - Expenses = Net Profit/Loss
        """
        # For Income Statement we only look at activity within the period
        results = db.query(
            models.GLAccount.account_code,
            models.GLAccount.account_name,
            models.GLAccount.account_type,
            func.sum(
                case((models.GLJournalEntry.entry_type == 'DEBIT', models.GLJournalEntry.amount), else_=0)
            ).label('total_debit'),
            func.sum(
                case((models.GLJournalEntry.entry_type == 'CREDIT', models.GLJournalEntry.amount), else_=0)
            ).label('total_credit')
        ).join(
            models.GLJournalEntry, 
            models.GLAccount.id == models.GLJournalEntry.gl_account_id
        ).filter(
            func.date(models.GLJournalEntry.entry_date) >= start_date,
            func.date(models.GLJournalEntry.entry_date) <= end_date,
            models.GLAccount.account_type.in_(['INCOME', 'EXPENSE'])
        ).group_by(
            models.GLAccount.account_code,
            models.GLAccount.account_name,
            models.GLAccount.account_type
        ).order_by(models.GLAccount.account_code).all()

        income = []
        expenses = []
        total_income = 0.0
        total_expenses = 0.0

        for row in results:
            debit = float(row.total_debit or 0)
            credit = float(row.total_credit or 0)
            
            if row.account_type == 'INCOME':
                balance = credit - debit
                income.append({
                    "account_code": row.account_code,
                    "account_name": row.account_name,
                    "balance": balance
                })
                total_income += balance
            elif row.account_type == 'EXPENSE':
                balance = debit - credit
                expenses.append({
                    "account_code": row.account_code,
                    "account_name": row.account_name,
                    "balance": balance
                })
                total_expenses += balance

        net_profit = total_income - total_expenses

        return {
            "period_start": start_date,
            "period_end": end_date,
            "income": {"items": income, "total": total_income},
            "expenses": {"items": expenses, "total": total_expenses},
            "net_profit": net_profit
        }

    @staticmethod
    def generate_par_report(db: Session, as_of_date: date, officer_id: Optional[int] = None) -> Dict[str, Any]:
        """
        Generates Portfolio At Risk (PAR) Report.
        Groups delinquent loans into 30, 60, 90+ day buckets.
        If officer_id is provided, only aggregates loans assigned to that officer.
        """
        # This queries the Loan table directly since it's a sub-ledger report
        # In a strict environment, we'd reconcile this sub-ledger total with the GL 1210 total
        
        query = db.query(models.Loan).filter(
            models.Loan.status.in_(['ACTIVE', 'DELINQUENT']),
            models.Loan.amount_outstanding > 0
        )
        
        if officer_id is not None:
             query = query.filter(models.Loan.applied_by == officer_id)
             
        loans = query.all()

        par_buckets = {
            "current": {"count": 0, "principal_outstanding": 0.0},
            "par_30": {"count": 0, "principal_outstanding": 0.0},
            "par_60": {"count": 0, "principal_outstanding": 0.0},
            "par_90_plus": {"count": 0, "principal_outstanding": 0.0}
        }
        
        total_portfolio = 0.0

        for loan in loans:
            outstanding = float(loan.amount_outstanding)
            total_portfolio += outstanding
            
            if loan.status == 'ACTIVE' or loan.delinquency_days == 0:
                par_buckets["current"]["count"] += 1
                par_buckets["current"]["principal_outstanding"] += outstanding
            elif loan.delinquency_days <= 30:
                par_buckets["par_30"]["count"] += 1
                par_buckets["par_30"]["principal_outstanding"] += outstanding
            elif loan.delinquency_days <= 60:
                par_buckets["par_60"]["count"] += 1
                par_buckets["par_60"]["principal_outstanding"] += outstanding
            else:
                par_buckets["par_90_plus"]["count"] += 1
                par_buckets["par_90_plus"]["principal_outstanding"] += outstanding

        par_ratio = 0.0
        if total_portfolio > 0:
             # PAR is typically defined as outstanding balance of loans > 30 days late
             total_at_risk = par_buckets["par_30"]["principal_outstanding"] + \
                             par_buckets["par_60"]["principal_outstanding"] + \
                             par_buckets["par_90_plus"]["principal_outstanding"]
             par_ratio = (total_at_risk / total_portfolio) * 100

        return {
            "as_of_date": as_of_date,
            "total_portfolio_outstanding": total_portfolio,
            "par_ratio_percentage": round(par_ratio, 2),
            "buckets": par_buckets
        }

    @staticmethod
    def generate_daily_cash_flow(db: Session, target_date: date) -> Dict[str, Any]:
        """
        Generates the Daily Cash Flow Matrix Report.
        Groups transactions by 'purpose' and pivots by 'payment_channel'.
        Supports structured categories for the Cameroonian standard layout.
        """
        from datetime import datetime
        start_of_day = datetime.combine(target_date, datetime.min.time())
        end_of_day = datetime.combine(target_date, datetime.max.time())

        # 1. Fetch Transactions for the day
        transactions = db.query(
            models.Transaction.purpose,
            models.Transaction.payment_channel,
            models.Transaction.transaction_type,
            func.sum(models.Transaction.amount).label('total_amount'),
            func.string_agg(models.Transaction.external_reference, ', ').label('refs'),
            func.string_agg(models.Transaction.comments, '; ').label('all_comments')
        ).filter(
            models.Transaction.created_at >= start_of_day,
            models.Transaction.created_at <= end_of_day
        ).group_by(
            models.Transaction.purpose,
            models.Transaction.payment_channel,
            models.Transaction.transaction_type
        ).all()

        # 2. Calculate Brought Forward (B/F) Balances per channel
        bf_query = db.query(
            models.Transaction.payment_channel,
            func.sum(
                case(
                    (models.Transaction.transaction_type.in_([
                        models.TransactionType.DEPOSIT, 
                        models.TransactionType.LOAN_REPAYMENT,
                        models.TransactionType.FEE,
                        models.TransactionType.INTEREST
                    ]), models.Transaction.amount),
                    (models.Transaction.transaction_type.in_([
                        models.TransactionType.WITHDRAWAL,
                        models.TransactionType.LOAN_DISBURSEMENT,
                        models.TransactionType.TRANSFER
                    ]), -models.Transaction.amount),
                    else_=0
                )
            ).label('balance')
        ).filter(
            models.Transaction.created_at < start_of_day
        ).group_by(
            models.Transaction.payment_channel
        ).all()

        bf_row = {
            "description": "BROUGHT FORWARD BALANCES",
            "corp_banks": 0.0, "mf_balico": 0.0, "mf_a": 0.0, "mf_glovic": 0.0, 
            "cash": 0.0, "mobile_om": 0.0, "mobile_mtn": 0.0, "total": 0.0, "refs": "", "comments": ""
        }
        
        for row in bf_query:
            amount = float(row.balance or 0)
            channel = row.payment_channel
            if channel == models.PaymentChannel.BANK_TRANSFER: bf_row["corp_banks"] += amount
            elif channel == models.PaymentChannel.BALI_CO: bf_row["mf_balico"] += amount
            elif channel == models.PaymentChannel.MICROFINANCE_A: bf_row["mf_a"] += amount
            elif channel == models.PaymentChannel.GLOVIC: bf_row["mf_glovic"] += amount
            elif channel == models.PaymentChannel.CASH: bf_row["cash"] += amount
            elif channel == models.PaymentChannel.ORANGE_MONEY: bf_row["mobile_om"] += amount
            elif channel == models.PaymentChannel.MTN_MOMO: bf_row["mobile_mtn"] += amount
            bf_row["total"] += amount

        # Helper to map TreasuryAccount to cash flow column
        def map_treasury_to_col(treasury_account):
            if not treasury_account: return "cash"
            if treasury_account.account_type == models.TreasuryAccountType.VAULT: return "cash"
            if treasury_account.account_type == models.TreasuryAccountType.BANK: return "corp_banks"
            if treasury_account.account_type == models.TreasuryAccountType.CREDIT_UNION: return "mf_balico"
            if treasury_account.account_type == models.TreasuryAccountType.MOBILE_MONEY:
                return "mobile_om" if "orange" in (treasury_account.name or "").lower() else "mobile_mtn"
            return "cash"

        # 2b. Add Brought Forward Balances from internal Treasury Transfers (VaultTransfers)
        bf_transfers = db.query(models.VaultTransfer).filter(
            models.VaultTransfer.status == models.VaultTransferStatus.APPROVED,
            models.VaultTransfer.approved_at < start_of_day
        ).all()
        
        # Pre-cache TreasuryAccounts to avoid N+1 queries in loops
        all_treasuries = {t.id: t for t in db.query(models.TreasuryAccount).all()}
        
        for transfer in bf_transfers:
            amount = float(transfer.amount)
            src_acc = all_treasuries.get(transfer.source_treasury_id) if transfer.source_treasury_id else None
            dst_acc = all_treasuries.get(transfer.destination_treasury_id) if transfer.destination_treasury_id else None
            
            src_col = map_treasury_to_col(src_acc) if transfer.source_treasury_id else "cash"
            dst_col = map_treasury_to_col(dst_acc) if transfer.destination_treasury_id else "cash"
            
            # Decrement source, increment destination
            bf_row[src_col] -= amount
            bf_row[dst_col] += amount
            # bf_row["total"] remains unchanged because it's an internal transfer

        # 3. Define Sections with sub-categories
        sections = {
            "INFLOWS": [],
            "EXPENSES_A": [],
            "EXPENSES_B": [],
            "PROJECTED_A": [],
            "PROJECTED_B": [],
            "PROJECTED_C": []
        }

        # Purpose mapping (Standardized)
        purpose_map = {
            # Inflows
            "CASH FROM BALICOP": "INFLOWS",
            "SAVINGS": "INFLOWS",
            "CURRENT ACT": "INFLOWS",
            "SHARE CAPITAL": "INFLOWS",
            "ENTRANCE FEES": "INFLOWS",
            "ENTRANCE FEE": "INFLOWS",
            "ACCOUNT OPENING FEE": "INFLOWS",
            "SOLIDARITY FUND": "INFLOWS",
            "BUILDING CONTRIBUTION": "INFLOWS",
            "LOAN REPAYMENT": "INFLOWS",
            "PREFERENCE SHARES": "INFLOWS",
            "LOAN PROCESSING FEES": "INFLOWS",
            
            # Expenses
            "PAYMENT OF A SUPPLIER": "EXPENSES_A",
            "CASH EXPENSES": "EXPENSES_A",
            "EXPENSES CLIENT": "EXPENSES_B",
            
            # Projected
            "SALARIES": "PROJECTED_A",
            "LOANS": "PROJECTED_B",
            "TAXATION AND CNPS": "PROJECTED_C"
        }

        INFLOW_TYPES = {
            models.TransactionType.DEPOSIT,
            models.TransactionType.LOAN_REPAYMENT,
            models.TransactionType.FEE,
            models.TransactionType.INTEREST,
            models.TransactionType.NJANGI_CONTRIBUTION,
        }
        OUTFLOW_TYPES = {
            models.TransactionType.WITHDRAWAL,
            models.TransactionType.LOAN_DISBURSEMENT,
            models.TransactionType.TRANSFER,
            models.TransactionType.NJANGI_PAYOUT,
        }

        rows_data = {}
        for row in transactions:
            p_orig = (row.purpose or "OTHER").replace("_", " ").upper()
            section = purpose_map.get(p_orig, "INFLOWS")
            
            if p_orig not in rows_data:
                rows_data[p_orig] = {
                    "description": p_orig,
                    "section": section,
                    "refs": "",
                    "corp_banks": 0.0, "mf_balico": 0.0, "mf_a": 0.0, "mf_glovic": 0.0,
                    "cash": 0.0, "mobile_om": 0.0, "mobile_mtn": 0.0, 
                    "total": 0.0, "comments": ""
                }
            
            # Use signed amount: inflow types are positive, outflow types are negative
            raw_amount = float(row.total_amount)
            if row.transaction_type in OUTFLOW_TYPES:
                amount = -raw_amount
            else:
                amount = raw_amount  # Default to positive for inflow types

            channel = row.payment_channel
            if channel == models.PaymentChannel.BANK_TRANSFER: rows_data[p_orig]["corp_banks"] += amount
            elif channel == models.PaymentChannel.BALI_CO: rows_data[p_orig]["mf_balico"] += amount
            elif channel == models.PaymentChannel.MICROFINANCE_A: rows_data[p_orig]["mf_a"] += amount
            elif channel == models.PaymentChannel.GLOVIC: rows_data[p_orig]["mf_glovic"] += amount
            elif channel == models.PaymentChannel.CASH: rows_data[p_orig]["cash"] += amount
            elif channel == models.PaymentChannel.ORANGE_MONEY: rows_data[p_orig]["mobile_om"] += amount
            elif channel == models.PaymentChannel.MTN_MOMO: rows_data[p_orig]["mobile_mtn"] += amount
            
            rows_data[p_orig]["total"] += amount
            if row.refs: rows_data[p_orig]["refs"] = (rows_data[p_orig]["refs"] + ", " + row.refs).strip(", ")
            if row.all_comments: rows_data[p_orig]["comments"] = (rows_data[p_orig]["comments"] + "; " + row.all_comments).strip("; ")

        # 3b. Add internal Treasury Transfers for the current day
        day_transfers = db.query(models.VaultTransfer).filter(
            models.VaultTransfer.status == models.VaultTransferStatus.APPROVED,
            models.VaultTransfer.approved_at >= start_of_day,
            models.VaultTransfer.approved_at <= end_of_day
        ).all()
        
        if day_transfers:
            if "TREASURY_TRANSFER" not in rows_data:
                rows_data["TREASURY_TRANSFER"] = {
                    "description": "INTERNAL TREASURY TRANSFERS",
                    "section": "INFLOWS",
                    "refs": "",
                    "corp_banks": 0.0, "mf_balico": 0.0, "mf_a": 0.0, "mf_glovic": 0.0,
                    "cash": 0.0, "mobile_om": 0.0, "mobile_mtn": 0.0, 
                    "total": 0.0, "comments": "System Treasury Sweeps"
                }
            
            for transfer in day_transfers:
                amount = float(transfer.amount)
                src_acc = all_treasuries.get(transfer.source_treasury_id) if transfer.source_treasury_id else None
                dst_acc = all_treasuries.get(transfer.destination_treasury_id) if transfer.destination_treasury_id else None
                
                src_col = map_treasury_to_col(src_acc) if transfer.source_treasury_id else "cash"
                dst_col = map_treasury_to_col(dst_acc) if transfer.destination_treasury_id else "cash"
                
                rows_data["TREASURY_TRANSFER"][src_col] -= amount
                rows_data["TREASURY_TRANSFER"][dst_col] += amount
                if transfer.transfer_ref:
                    rows_data["TREASURY_TRANSFER"]["refs"] += f"{transfer.transfer_ref}, "
            
            # Clean up trailing comma
            rows_data["TREASURY_TRANSFER"]["refs"] = rows_data["TREASURY_TRANSFER"]["refs"].rstrip(", ")

        for data in rows_data.values():
            sections[data["section"]].append(data)

        # 4. Calculate cumulative sub-totals specifically for the "TOTAL CASH AND BANK INFLOWS" row
        # In the requested layout, this row is (Brought Forward + Inflows)
        cumulative_inflows = {k: bf_row[k] for k in ["corp_banks", "mf_balico", "mf_a", "mf_glovic", "cash", "mobile_om", "mobile_mtn", "total"]}
        for row in sections["INFLOWS"]:
            for k in cumulative_inflows:
                cumulative_inflows[k] += row.get(k, 0.0)
        
        cumulative_inflows["description"] = "TOTAL CASH AND BANK INFLOWS"

        return {
            "date": target_date.strftime("%d/%m/%Y"),
            "year": target_date.year,
            "version": "v2.2-MATRIX",
            "brought_forward": bf_row,
            "cumulative_inflows": cumulative_inflows,
            "sections": sections
        }

    @staticmethod
    def export_to_excel(report_name: str, data: List[Dict[str, Any]] | Dict[str, Any]) -> str:
        """
        Exports report data to an Excel file and returns the file path.
        Handles specialized formatting for Daily Cash Flow reports.
        """
        import pandas as pd
        import tempfile
        import os

        logger.info(f"EXCEL EXPORT: report_name={report_name}, type(data)={type(data)}")

        # Specialized formatting for Daily Cash Flow
        if report_name == "daily_cash_flow":
            logger.info("DASHBOARD: STARTING SPECIALIZED DAILY CASH FLOW EXCEL FORMATTER")
            if not isinstance(data, dict) or "sections" not in data:
                logger.error(f"CRITICAL: Daily Cash Flow data is invalid format for specialized exporter. Type: {type(data)}")
                # Proceed anyway to see if it can handle it, or it will fail naturally
            return ReportingService._format_daily_cash_flow_excel(data)

        logger.info("FALLING BACK TO GENERIC PANDAS EXCEL EXPORT")

        # Specialized formatter for Trial Balance
        if report_name == "trial_balance" and isinstance(data, dict) and "rows" in data:
            return ReportingService._format_trial_balance_excel(data)

        # Handle different data structures for generic export
        if isinstance(data, dict):
            if "items" in data:
                df = pd.DataFrame(data["items"])
            elif "rows" in data:
                df = pd.DataFrame(data["rows"])
            else:
                df = pd.DataFrame([data])
        else:
            df = pd.DataFrame(data)

        # Create temp file
        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix=f"{report_name}_")
        os.close(fd)
        
        df.to_excel(path, index=False)
        return path

    @staticmethod
    def _format_trial_balance_excel(data: Dict[str, Any]) -> str:
        """
        Produces a formatted Excel Trial Balance with header, account rows grouped by type,
        a bold totals row at the bottom, and a balanced/unbalanced indicator.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        import tempfile, os

        wb = Workbook()
        ws = wb.active
        ws.title = "Trial Balance"

        blue_fill = PatternFill(start_color="0070C0", fill_type="solid")
        light_blue_fill = PatternFill(start_color="BDD7EE", fill_type="solid")
        total_fill = PatternFill(start_color="1F5C99", fill_type="solid")
        white_font = Font(bold=True, color="FFFFFF", size=10)
        bold_font = Font(bold=True, size=10)
        normal_font = Font(size=10)
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right")
        thin = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        num_fmt = '#,##0.00'

        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"TRIAL BALANCE (Balance Générale) — As of {data.get('as_of_date', '')}"
        ws['A1'].font = Font(bold=True, color="FFFFFF", size=13)
        ws['A1'].fill = blue_fill
        ws['A1'].alignment = center

        # Header row
        headers = ["Account Code", "Account Name", "Account Type", "Opening Balance", "Debit (Dr)", "Credit (Cr)", "Closing Balance"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = white_font
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[3].height = 22

        # Data rows
        current_row = 4
        type_colors = {
            "ASSET": "E2EFDA", "LIABILITY": "FCE4D6", "EQUITY": "EAD1DC",
            "INCOME": "D9EAD3", "EXPENSE": "FFF2CC"
        }

        for row in data.get("rows", []):
            acc_type = (row.get("account_type") or "").upper()
            fill_color = type_colors.get(acc_type)
            row_fill = PatternFill(start_color=fill_color, fill_type="solid") if fill_color else None
            values = [
                row.get("account_code", ""),
                row.get("account_name", ""),
                acc_type,
                row.get("opening_balance", 0),
                row.get("debit", 0),
                row.get("credit", 0),
                row.get("closing_balance", 0),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.font = normal_font
                cell.border = thin
                if isinstance(val, float) or isinstance(val, int) and col > 3:
                    cell.number_format = num_fmt
                    cell.alignment = right
                if row_fill:
                    cell.fill = row_fill
            current_row += 1

        # Totals row
        totals = data.get("totals", {})
        total_values = [
            "", "TOTALS", "",
            totals.get("opening_balance", 0),
            totals.get("debit", 0),
            totals.get("credit", 0),
            totals.get("net_balance", 0),
        ]
        for col, val in enumerate(total_values, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = total_fill
            cell.border = thin
            if isinstance(val, (float, int)) and col > 3:
                cell.number_format = num_fmt
                cell.alignment = right

        # Balance indicator
        current_row += 1
        is_balanced = totals.get("is_balanced", False)
        status_text = "✔ BALANCED — Debits equal Credits" if is_balanced else "✘ UNBALANCED — Check GL Entries"
        status_cell = ws.cell(row=current_row, column=1, value=status_text)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        status_cell.font = Font(bold=True, color="FFFFFF", size=10)
        status_cell.fill = PatternFill(start_color="375623" if is_balanced else "C00000", fill_type="solid")
        status_cell.alignment = center

        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 14
        for col_letter in 'DEFG':
            ws.column_dimensions[col_letter].width = 20

        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="TrialBalance_")
        os.close(fd)
        wb.save(path)
        return path


    @staticmethod
    def _format_daily_cash_flow_excel(report_data: Dict[str, Any]) -> str:
        """
        Advanced Excel formatting for Daily Cash Flow Statement matching the requested image.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        import tempfile
        import os

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Cash Flow"

        # Styles
        title_font = Font(bold=True, color="FFFFFF", size=14)
        header_font = Font(bold=True, color="FFFFFF", size=9)
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        blue_fill = PatternFill(start_color="0070C0", fill_type="solid") # Stronger Blue
        light_blue_fill = PatternFill(start_color="BDD7EE", fill_type="solid") # Light Blue for sub-totals
        grey_fill = PatternFill(start_color="D9D9D9", fill_type="solid") # Standard Grey for headers
        
        # 1. Title
        ws.merge_cells('A1:L1')
        title_cell = ws['A1']
        title_cell.value = f"DAILY CASH FLOW STATEMENT {report_data.get('year', 2026)} ({report_data.get('version', 'FINAL')})"
        title_cell.font = title_font
        title_cell.fill = blue_fill
        title_cell.alignment = center_align

        # 2. Main Headers (Rows 3-5)
        headers = [
            (1, "DATES"), (2, "DESCRIPTION"), (3, "REFS"), (11, "TOTAL"), (12, "COMMENTS")
        ]
        for col, text in headers:
            ws.merge_cells(start_row=3, start_column=col, end_row=5, end_column=col)
            cell = ws.cell(row=3, column=col)
            cell.value = text
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = center_align
            for r in range(3, 6): ws.cell(row=r, column=col).border = thin_border

        # Top level merge: PAYMENT CENTRES
        ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=10)
        cell = ws.cell(row=3, column=4)
        cell.value = "PAYMENT CENTRES"
        cell.font = header_font
        cell.fill = blue_fill
        cell.alignment = center_align
        for c in range(4, 11): ws.cell(row=3, column=c).border = thin_border

        # Row 4: Column groups
        groups = [
            (4, 4, "CORPORATE BANKS"), 
            (5, 7, "MICRO FINANCE INSTITUTIONS AND COOPERATIVES"),
            (8, 8, "CASH"),
            (9, 10, "MOBILE CASH")
        ]
        for start_c, end_c, text in groups:
            if start_c != end_c: ws.merge_cells(start_row=4, start_column=start_c, end_row=4, end_column=end_c)
            cell = ws.cell(row=4, column=start_c)
            cell.value = text
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = center_align
            for c in range(start_c, end_c + 1): ws.cell(row=4, column=c).border = thin_border

        # Row 5: Specific columns
        cols = [
            (4, "A"), (5, "BALI CO"), (6, "A"), (7, "GLOVIC"), (8, "CASH"), (9, "OM"), (10, "MTN MoMo")
        ]
        for col, text in cols:
            cell = ws.cell(row=5, column=col)
            cell.value = text
            cell.font = header_font
            cell.fill = blue_fill
            cell.alignment = center_align
            cell.border = thin_border

        # 3. Data Printing Logic
        current_row = 6
        
        def write_row(row_data, is_total=False, fill=None):
            nonlocal current_row
            ws.cell(row=current_row, column=1, value=report_data.get("date", ""))
            ws.cell(row=current_row, column=2, value=row_data.get("description", ""))
            ws.cell(row=current_row, column=3, value=row_data.get("refs", ""))
            
            mapping = {4: "corp_banks", 5: "mf_balico", 6: "mf_a", 7: "mf_glovic", 8: "cash", 9: "mobile_om", 10: "mobile_mtn", 11: "total"}
            for col, key in mapping.items():
                val = row_data.get(key, 0.0)
                cell = ws.cell(row=current_row, column=col, value=val if val != 0 else "-")
                if val != 0: cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                cell.border = thin_border
                if is_total: cell.font = bold_font
                if fill: cell.fill = fill

            for col in [1, 2, 3, 12]:
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                if is_total: cell.font = bold_font
                if fill: cell.fill = fill
            
            current_row += 1
            return row_data

        # Actual categories
        def sum_section(section_name):
            total = {k: 0.0 for k in ["corp_banks", "mf_balico", "mf_a", "mf_glovic", "cash", "mobile_om", "mobile_mtn", "total"]}
            for row in report_data["sections"].get(section_name, []):
                write_row(row)
                for k in total: total[k] += row.get(k, 0.0)
            return total

        # A. Brought Forward
        write_row(report_data["brought_forward"], fill=light_blue_fill)

        # B. Inflows
        _ = sum_section("INFLOWS") # Write individual inflow rows, ignore raw sum
        
        # Total Inflows Row (CUMULATIVE in the standard layout)
        inflow_total = report_data.get("cumulative_inflows", {})
        write_row(inflow_total, is_total=True, fill=light_blue_fill)

        # C. Expenses
        current_row += 1
        ws.cell(row=current_row, column=2, value="EXPENSES").font = bold_font
        current_row += 1
        
        ws.cell(row=current_row, column=2, value="A) CASH EXPENSES (office)")
        current_row += 1
        exp_a = sum_section("EXPENSES_A")
        exp_a["description"] = "TOTAL CASH EXPENSES"
        write_row(exp_a, is_total=True)

        current_row += 1
        ws.cell(row=current_row, column=2, value="B) EXPENSES CLIENT")
        current_row += 1
        exp_b = sum_section("EXPENSES_B")
        exp_b["description"] = "TOTAL CLIENTS EXPENSES"
        write_row(exp_b, is_total=True)

        full_exp = {k: exp_a[k] + exp_b[k] for k in exp_a if k != "description"}
        
        # D. Projected
        current_row += 1
        ws.cell(row=current_row, column=2, value="PROJECTED EXPENSES").font = bold_font
        current_row += 1

        ws.cell(row=current_row, column=2, value="A) SALARIES")
        current_row += 1
        proj_a = sum_section("PROJECTED_A")
        proj_a["description"] = "TOTAL SALARIES"
        write_row(proj_a, is_total=True)

        ws.cell(row=current_row, column=2, value="B) LOANS")
        current_row += 1
        proj_b = sum_section("PROJECTED_B")
        proj_b["description"] = "TOTAL LOANS"
        write_row(proj_b, is_total=True)

        ws.cell(row=current_row, column=2, value="D) TAXATION AND CNPS")
        current_row += 1
        proj_c = sum_section("PROJECTED_C")
        proj_c["description"] = "TOTAL TAXATION AND CNPS"
        write_row(proj_c, is_total=True)

        full_proj = {k: proj_a[k] + proj_b[k] + proj_c[k] for k in proj_a if k != "description"}
        full_proj["description"] = "TOTAL PROJECTIONS"
        write_row(full_proj, is_total=True)

        # Totals
        total_ex_pr = {k: full_exp[k] + full_proj[k] for k in full_exp}
        total_ex_pr["description"] = "TOTAL DAILY EXPS + PROJECTIONS"
        write_row(total_ex_pr, is_total=True, fill=light_blue_fill)

        # Balance @ Hand
        # Note: inflow_total (cumulative_inflows) already includes Brought Forward + day's Inflows.
        # So we must NOT add brought_forward again — it would double-count it.
        # Formula: BALANCE @ HAND = (BF + Inflows) - Cash Expenses - Client Expenses
        closing_bal = {
            "description": "TOTAL BALANCE @ HAND",
            "corp_banks": inflow_total["corp_banks"] - exp_a["corp_banks"] - exp_b["corp_banks"],
            "mf_balico": inflow_total["mf_balico"] - exp_a["mf_balico"] - exp_b["mf_balico"],
            "mf_a": inflow_total["mf_a"] - exp_a["mf_a"] - exp_b["mf_a"],
            "mf_glovic": inflow_total["mf_glovic"] - exp_a["mf_glovic"] - exp_b["mf_glovic"],
            "cash": inflow_total["cash"] - exp_a["cash"] - exp_b["cash"],
            "mobile_om": inflow_total["mobile_om"] - exp_a["mobile_om"] - exp_b["mobile_om"],
            "mobile_mtn": inflow_total["mobile_mtn"] - exp_a["mobile_mtn"] - exp_b["mobile_mtn"]
        }
        closing_bal["total"] = sum(v for k, v in closing_bal.items() if k != "description")
        write_row(closing_bal, is_total=True, fill=blue_fill)
        for c in range(1, 13): ws.cell(row=current_row-1, column=c).font = Font(bold=True, color="FFFFFF")

        # Balance After projections
        final_bal = {k: v for k, v in closing_bal.items()}
        for k in final_bal:
            if k != "description" and k in full_proj: final_bal[k] -= full_proj[k]
        final_bal["description"] = "BALANCE AFTER PROJECTIONS"
        # IMPORTANT: final_bal was copied from closing_bal which already has a 'total' key.
        # We must exclude 'total' from the sum, otherwise it gets counted twice.
        final_bal["total"] = sum(v for k, v in final_bal.items() if k not in ("description", "total"))
        write_row(final_bal, is_total=True, fill=blue_fill)
        for c in range(1, 13): ws.cell(row=current_row-1, column=c).font = Font(bold=True, color="FFFFFF")

        # Dimensions
        ws.column_dimensions['B'].width = 40
        for col_letter in 'CDEFGHIJKL': ws.column_dimensions[col_letter].width = 15

        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="DailyCashFlow_")
        os.close(fd)
        wb.save(path)
        return path

    @staticmethod
    def export_to_pdf(report_name: str, title: str, data: List[Dict[str, Any]]) -> str:
        """
        Exports list-based report data to a simple PDF table using fpdf2.
        """
        from fpdf import FPDF
        import tempfile
        import os

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('helvetica', 'B', 16)
        pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT", align='C')
        pdf.ln(10)

        if not data:
            pdf.set_font('helvetica', '', 12)
            pdf.cell(0, 10, "No data available.", new_x="LMARGIN", new_y="NEXT")
        else:
            # Table Header
            pdf.set_font('helvetica', 'B', 10)
            headers = list(data[0].keys())
            
            # Very basic column width division
            col_width = pdf.epw / len(headers)
            
            for header in headers:
                pdf.cell(col_width, 10, header.replace("_", " ").title(), border=1)
            pdf.ln()

            # Table Data
            pdf.set_font('helvetica', '', 8)
            grand_total_debit = 0.0
            grand_total_credit = 0.0

            for row in data:
                # Track max height for the row
                max_h = 10
                
                # Calculate grand totals if this is a Trial Balance
                if "debit" in row: grand_total_debit += float(row.get("debit", 0))
                if "credit" in row: grand_total_credit += float(row.get("credit", 0))

                # Start drawing cells
                for key in headers:
                    val = row.get(key, "")
                    if isinstance(val, float):
                        val = f"{val:,.0f}"
                    
                    # Implementation of text wrap using multi_cell
                    # Remember current position to draw borders correctly
                    x = pdf.get_x()
                    y = pdf.get_y()
                    pdf.multi_cell(col_width, 5, str(val), border=1, align='L')
                    # Update max height for the row move to the next cell's start position
                    pdf.set_xy(x + col_width, y)
                pdf.ln(10) # Move to next line after all columns done

            # Grand Totals Row
            if "debit" in data[0] or "credit" in data[0]:
                pdf.set_font('helvetica', 'B', 9)
                pdf.set_fill_color(240, 243, 246)
                pdf.cell(col_width * 3, 10, "GRAND TOTAL:", border=1, fill=True, align='R')
                pdf.cell(col_width, 10, f"{grand_total_debit:,.0f}", border=1, fill=True)
                pdf.cell(col_width, 10, f"{grand_total_credit:,.0f}", border=1, fill=True)
                # Fill remaining cells
                for _ in range(len(headers) - 5):
                    pdf.cell(col_width, 10, "", border=1, fill=True)
                pdf.ln()

        # Create temp file
        fd, path = tempfile.mkstemp(suffix=".pdf", prefix=f"{report_name}_")
        os.close(fd)
        
        pdf.output(path)
        return path

    @staticmethod
    def generate_cobac_liquidity(db: Session, period: str = 'daily') -> Dict[str, Any]:
        """
        Calculates the COBAC Liquidity Ratio.
        Ratio = (Liquid Assets) / (Short-term Liabilities)
        Standard requires > 100% for compliance.
        """
        # 1. Liquid Assets (Cash + Banks + Equivalents - Class 1, Category 10)
        # We sum all journal entries for accounts starting with '10'
        liquid_assets_query = db.query(
            func.sum(
                case(
                    (models.GLJournalEntry.entry_type == 'DEBIT', models.GLJournalEntry.amount),
                    else_=-models.GLJournalEntry.amount
                )
            )
        ).join(
            models.GLAccount, 
            models.GLAccount.id == models.GLJournalEntry.gl_account_id
        ).filter(
            models.GLAccount.account_code.like('10%')
        ).scalar() or 0.0

        # 2. Short-term Liabilities (Member Deposits - Class 2, Category 20)
        # We sum all journal entries for accounts starting with '20'
        # For liabilities, Credits increase the balance
        liabilities_query = db.query(
            func.sum(
                case(
                    (models.GLJournalEntry.entry_type == 'CREDIT', models.GLJournalEntry.amount),
                    else_=-models.GLJournalEntry.amount
                )
            )
        ).join(
            models.GLAccount, 
            models.GLAccount.id == models.GLJournalEntry.gl_account_id
        ).filter(
            models.GLAccount.account_code.like('20%')
        ).scalar() or 0.0

        liquid_assets = float(liquid_assets_query)
        short_term_liabilities = float(liabilities_query)

        # Calculate ratio
        if short_term_liabilities > 0:
            ratio = (liquid_assets / short_term_liabilities) * 100
        else:
            # If no liabilities but we have cash, we are effectively 100%+ liquid
            ratio = 150.0 if liquid_assets > 0 else 100.0
            
        status = "COMPLIANT" if ratio >= 100 else "NON-COMPLIANT"
        
        return {
            "ratio": round(ratio, 2),
            "liquidity_ratio": round(ratio, 2), # Compatibility with Director/Board Dashboard
            "status": status,
            "liquid_assets": liquid_assets,
            "short_term_liabilities": short_term_liabilities,
            "period": period,
            "as_of": date.today().isoformat()
        }

    @staticmethod
    def get_board_metrics(db: Session) -> Dict[str, Any]:
        """
        Aggregates data for the Board Executive Overview.
        Includes Sector Distribution, Branch Performance, and Critical Anomalies.
        """
        # 1. Loan Distribution by "Sector" (using Product Name as proxy)
        sector_dist = db.query(
            models.LoanProduct.name,
            func.sum(models.Loan.amount_outstanding).label('value')
        ).join(models.Loan, models.Loan.product_id == models.LoanProduct.id) \
         .filter(models.Loan.status.in_(['ACTIVE', 'DELINQUENT'])) \
         .group_by(models.LoanProduct.name).all()
        
        sector_data = [{"name": row.name, "value": float(row.value or 0)} for row in sector_dist]

        # 2. Branch Performance (Deposits vs Loans)
        branches = db.query(models.Branch).all()
        branch_data = []
        for b in branches:
            loans = db.query(func.sum(models.Loan.amount_outstanding)) \
                      .join(models.Member, models.Loan.member_id == models.Member.id) \
                      .filter(models.Member.branch_id == b.id).scalar() or 0
            deposits = db.query(func.sum(models.Account.balance)) \
                         .join(models.Member, models.Account.member_id == models.Member.id) \
                         .filter(models.Member.branch_id == b.id).scalar() or 0
            branch_data.append({"name": b.name, "loans": float(loans), "deposits": float(deposits)})

        # 3. Anomaly Radar (Latest Audit Logs for Security Events)
        anomalies_raw = db.query(models.AuditLog).filter(
            models.AuditLog.action.in_(['LOAN_DISBURSEMENT', 'USER_LOGIN_FAILED', 'PERMISSION_DENIED', 'LARGE_TRANSACTION', 'SYSTEM_OVERRIDE'])
        ).order_by(models.AuditLog.created_at.desc()).limit(5).all()
        
        anomalies = []
        for a in anomalies_raw:
            severity = "CRITICAL" if a.action in ['PERMISSION_DENIED', 'SYSTEM_OVERRIDE'] else "WARNING"
            anomalies.append({
                "id": a.id,
                "type": severity,
                "msg": a.description or a.action,
                "time": a.created_at.strftime("%I:%M %p"),
                "location": "System" # Could derive from IP/User branch
            })

        return {
            "sector_data": sector_data,
            "branch_data": branch_data,
            "anomalies": anomalies
        }

    @staticmethod
    def get_loan_dossier(db: Session, loan_id: int) -> Dict[str, Any]:
        """
        Aggregates a full dossier for the Credit Committee review.
        Includes Scorecard data (Tenure, Savings, History).
        """
        loan = db.query(models.Loan).filter(models.Loan.id == loan_id).first()
        if not loan:
            return {}
            
        member = loan.member
        # Tenure in months
        from datetime import datetime
        tenure_days = (datetime.now() - member.created_at).days
        tenure_months = tenure_days // 30

        # Total Savings across all accounts
        total_savings = db.query(func.sum(models.Account.balance)).filter(
            models.Account.member_id == member.id,
            models.Account.account_type.in_(['SAVINGS', 'SHARES'])
        ).scalar() or 0

        # Loan History
        past_loans = db.query(models.Loan).filter(
            models.Loan.member_id == member.id,
            models.Loan.id != loan_id
        ).all()
        
        history = {
            "total_count": len(past_loans),
            "fully_repaid": len([l for l in past_loans if l.status == 'CLOSED']),
            "total_borrowed": float(sum(l.principal_amount for l in past_loans)),
            "delinquency_history": any(l.delinquency_days > 0 for l in past_loans)
        }

        return {
            "loan_id": loan.id,
            "member_tenure_months": tenure_months,
            "total_savings": float(total_savings),
            "loan_history": history,
            "is_insider": loan.is_insider_loan,
            "collateral": [
                {"type": "Savings", "value": float(total_savings), "description": "Lien on shares/savings"},
                # ... other collateral logic ...
            ]
        }
